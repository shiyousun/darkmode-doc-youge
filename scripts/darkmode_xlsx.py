#!/usr/bin/env python3
"""
darkmode_xlsx.py — 把 Excel (.xlsx) 转成护眼模式

处理项：
- 所有有内容的单元格：主题底色 + 主题前景色字体
- 所有边框改主题边框色
- 工作表标签底色改主题底色

用法：
    python3 darkmode_xlsx.py input.xlsx [output.xlsx] [--theme warm]
"""
from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Color, Font, PatternFill, Side

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from themes import DEFAULT_THEME, THEMES, get_theme


def _argb(hex6: str) -> str:
    """openpyxl 颜色格式是 8 位 aRGB，需要补上 FF 透明度。"""
    h = hex6.upper().lstrip("#")
    if len(h) == 8:
        return h
    return "FF" + h


def darken_font(font: Font, fg_argb: str) -> Font:
    new_color = Color(rgb=fg_argb)
    return Font(
        name=font.name,
        size=font.size,
        bold=font.bold,
        italic=font.italic,
        vertAlign=font.vertAlign,
        underline=font.underline,
        strike=font.strike,
        color=new_color,
        family=font.family,
        scheme=font.scheme,
        charset=font.charset,
    )


def convert_xlsx_to_darkmode(
    input_path: Path,
    output_path: Path,
    theme_name: str = DEFAULT_THEME,
) -> dict:
    theme = get_theme(theme_name)
    bg_argb = _argb(theme["bg"])
    fg_argb = _argb(theme["fg"])
    border_argb = _argb(theme["border"])
    tab_color = theme["bg"]  # tabColor 用 6 位 hex

    border_side = Side(style="thin", color=border_argb)
    border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )
    fill = PatternFill(start_color=bg_argb, end_color=bg_argb, fill_type="solid")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(input_path), str(output_path))

    wb = load_workbook(str(output_path))
    sheet_count = 0
    cell_count = 0

    for ws in wb.worksheets:
        sheet_count += 1

        try:
            ws.sheet_properties.tabColor = tab_color
        except Exception:
            pass

        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = fill
                cell.font = darken_font(cell.font or Font(), fg_argb)
                cell.border = border
                cell_count += 1

        for merged_range in list(ws.merged_cells.ranges):
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            top_left.fill = fill
            top_left.font = darken_font(top_left.font or Font(), fg_argb)
            top_left.border = border

    wb.save(str(output_path))

    return {
        "sheets": sheet_count,
        "cells": cell_count,
        "theme": theme_name,
        "size_in": input_path.stat().st_size,
        "size_out": output_path.stat().st_size,
    }


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="XLSX 护眼模式转换器")
    parser.add_argument("input", help="输入 .xlsx 路径")
    parser.add_argument("output", nargs="?", help="输出路径（默认 xxx_dark.xlsx）")
    parser.add_argument(
        "--theme",
        default=DEFAULT_THEME,
        choices=list(THEMES.keys()),
        help=f"护眼主题（默认 {DEFAULT_THEME}）",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        print(f"❌ 输入文件不存在: {input_path}", file=sys.stderr)
        return 2

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        output_path = input_path.with_name(f"{input_path.stem}_dark.xlsx")

    theme = get_theme(args.theme)
    print(f"▶ XLSX 护眼模式转换")
    print(f"  输入: {input_path}")
    print(f"  输出: {output_path}")
    print(f"  主题: {args.theme} ({theme['name']}) bg=#{theme['bg']} fg=#{theme['fg']}")

    info = convert_xlsx_to_darkmode(input_path, output_path, theme_name=args.theme)
    print(
        f"✅ 完成 · {info['sheets']} 工作表 · 处理 {info['cells']} 单元格 · "
        f"{info['size_in'] / 1024:.1f} KB → {info['size_out'] / 1024:.1f} KB"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
